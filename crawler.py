import requests
from bs4 import BeautifulSoup as BS
from collections import OrderedDict
from openpyxl import Workbook, styles
from math import ceil
import threading
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry

def requests_retry_session(
    retries=10,
    backoff_factor=1,
    status_forcelist=(500, 502, 504),
    session=None,
):
    session = session or requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session

def export_excel(kinders, header):
    wb = Workbook()
    ws = wb.active

    sheet = wb['Sheet']
    widths = [13, 15, 35, 16, 16, 9, 9, 9, 18]
    sheet.row_dimensions[1].height = 30
    for i, h in enumerate(header):
        sheet.column_dimensions[chr(65 + i)].width = widths[i]
        cell = ws.cell(row=1, column=i + 1)
        cell.value = h
        cell.alignment = styles.Alignment(horizontal='center', vertical='center')

    for row, data in enumerate(kinders, start=2):
        for col, v in enumerate(data.values(), start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = v
            cell.alignment = styles.Alignment(horizontal='center')

    wb.save('result.xlsx')


def getData(kinders, ittId, err):

    session = requests_retry_session()

    data = {'ittId': ittId}

    res = session.post("https://e-childschoolinfo.moe.go.kr/kinderMt/kinderSummary.do",
                        data=data).text
    summary = BS(res, 'lxml')

    title = summary.select(
        '#rightContent > div.pageBox > div > div.content > table > tbody > tr:nth-child(1) > td:nth-child(2)')[
        0].text.strip()
    s1, s2 = summary.select('#rightContent > div.pageBox > div > div.content > table > tbody > tr:nth-child(7) > td')[
        0].text.strip().split('/')

    res = session.post("https://e-childschoolinfo.moe.go.kr/kinderMt/kinderRevAndExp.do",
                        data=data).text
    
    spends = OrderedDict([('교육청명', s1.strip()),
                          ('교육지원청명', s2.strip()),
                          ('유치원명', title),
                          ('교사연수ㆍ연구비', ''),
                          ('교재ㆍ교구구입비', ''),
                          ('행사비', ''),
                          ('장학금', ''),
                          ('복리비', ''),
                          ('일반급식비ㆍ간식비', '')])
    soup = BS(res, 'lxml')

    try:
        n = 19
        for i in range(19, 22):
            p = soup.select(f'#miniTab4 > table > tbody > tr:nth-child({i})')[0].find_all('td')[
                1].text.strip()
            if p == '일반교육활동비':        
                n = i
                break

        for key in list(spends.keys())[3:]:
            spends[key] = soup.select(f'#miniTab4 > table > tbody > tr:nth-child({n})')[0].find_all('td')[
                -2].text.strip()
            n += 1
        kinders.append(spends)
    except:
        err.append(title)


def f():
    session = requests_retry_session()
    data = {'tabNum': '3', 'pageIndex': "1", 'kinderEstablishType': '97'}

    res = session.post("https://e-childschoolinfo.moe.go.kr/kinderMt/combineFind.do",
                        data=data).text
    pages = BS(res, 'lxml')

    length = int(pages.select('#rightContent > div.pageBox > div > div.content > p.tblResult > span')[0].text.replace(',',''))

    err = []
    kinders = []

    pag = 10

    for page in range(1, ceil(length / 10) + 1):

        # 마지막 페이지
        if page == ceil(length / 10):
            pag = length%10 or 10
        for i in range(pag):
            data = {'tabNum': '3', 'pageIndex': page, 'kinderEstablishType': '97'}
            res = session.post("https://e-childschoolinfo.moe.go.kr/kinderMt/combineFind.do", data=data).text
            pages = BS(res, 'lxml')
            try:
                ittId = pages.find('input', {'id': f'kinderCompare{i}'})['value'][:36]
                t = threading.Thread(target=getData, args=[kinders, ittId, err])
                t.start()
            except:
                print(page, pag, res)
        print(page, '/', ceil(length/10))
    for thread in threading.enumerate():
        if thread.daemon:
            continue
        try:
            thread.join()
        except RuntimeError as err:
            if 'cannot join current thread' in err.args[0]:
                # catchs main thread
                continue
            else:
                raise

    export_excel(kinders, list(kinders[0].keys()))
    print('result.xlsx에 저장되었습니다.')

    input()


f()
                 
